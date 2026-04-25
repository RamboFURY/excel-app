from PyQt6.QtCore import QThread, pyqtSignal
from processor import process_file
from openpyxl import load_workbook
from datetime import datetime
import os
import pandas as pd


class Worker(QThread):
    progress = pyqtSignal(int)
    done = pyqtSignal()

    def __init__(self, files, year_map):
        super().__init__()
        self.files = files
        self.year_map = year_map

        # Save in project folder
        self.output_dir = os.path.dirname(os.path.abspath(__file__))
        print("📁 Saving to:", self.output_dir)

    def apply_formula(self, wb):
        year_map = self.year_map

        def apply(ws):
            headers = {}

            for col in range(1, ws.max_column + 1):
                val = ws.cell(row=1, column=col).value
                if val:
                    headers[val.strip()] = col

            if "Amount transferred" not in headers or "No. of Shares" not in headers:
                return

            amount_col = headers["Amount transferred"]
            shares_col = headers["No. of Shares"]
            year_col = headers.get("Proposed Date of transfer to IEPF(DD-MON-YYYY)")

            if year_col is None:
                return

            for row in range(2, ws.max_row + 1):
                amount_cell = ws.cell(row=row, column=amount_col).coordinate
                val = ws.cell(row=row, column=year_col).value

                year = None
                if isinstance(val, datetime):
                    year = val.year
                elif isinstance(val, int):
                    year = val

                if year not in year_map:
                    continue

                n1, n2 = year_map[year]
                formula = f"=ROUND(({amount_cell}/{n1})*{n2},0)"
                ws.cell(row=row, column=shares_col).value = formula

        for sheet in wb.sheetnames:
            apply(wb[sheet])

    def run(self):
        total = len(self.files)

        for i, f in enumerate(self.files):
            try:
                print(f"📂 Processing: {f}")

                df, df_folio, df_dpid, summary = process_file(f)

                filename = os.path.basename(f)
                output_path = os.path.join(self.output_dir, f"processed_{filename}")

                with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name="Data", index=False)
                    summary.to_excel(writer, sheet_name="Summary", index=False)
                    df_folio.to_excel(writer, sheet_name="Unique_Folio", index=False)
                    df_dpid.to_excel(writer, sheet_name="Unique_DPID", index=False)

                wb = load_workbook(output_path)
                self.apply_formula(wb)
                wb.save(output_path)

                print(f"✅ Saved: {output_path}")

            except Exception as e:
                print(f"❌ Error in {f}: {e}")

            percent = int((i + 1) / total * 100)
            self.progress.emit(percent)

        self.done.emit()